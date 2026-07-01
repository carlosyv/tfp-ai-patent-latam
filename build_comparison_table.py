#!/usr/bin/env python3
"""
build_comparison_table.py
=========================
Reads regression_results.json from v3 (with imputation) and v4 (no imputation)
and produces a publication-ready side-by-side comparison spreadsheet covering
H1–H5 and the Pesaran CD test.

Output: output/results/comparison_v3_vs_v4_noimput.xlsx
"""

import json
import numpy as np
import pandas as pd
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE  = Path(__file__).resolve().parent
V3    = BASE / 'output/results/benchmark_dissertation_v3/regression_results.json'
V4    = BASE / 'output/results/benchmark_dissertation_v4_noimput/regression_results.json'
OUT   = BASE / 'output/results/comparison_v3_vs_v4_noimput.xlsx'

with open(V3) as f: v3 = json.load(f)
with open(V4) as f: v4 = json.load(f)

# ── Helpers ────────────────────────────────────────────────────────────────────
def stars(p):
    if p is None or (isinstance(p, float) and np.isnan(p)): return ''
    if p < 0.01:  return '***'
    if p < 0.05:  return '**'
    if p < 0.10:  return '*'
    return ''

def fmt_coef(b, p):
    if b is None or (isinstance(b, float) and np.isnan(b)): return '—'
    return f'{b:.4f}{stars(p)}'

def fmt_p(p):
    if p is None or (isinstance(p, float) and np.isnan(p)): return '—'
    return f'{p:.3f}'

def delta_str(b3, b4):
    """Arrow showing direction and magnitude of change in coefficient."""
    if b3 is None or b4 is None: return '—'
    d = b4 - b3
    if abs(d) < 1e-6: return '≈'
    return f'{"↑" if d > 0 else "↓"} {d:+.4f}'

def delta_sig(p3, p4):
    """Signal whether significance improved / worsened / stable."""
    s3 = stars(p3) or 'NS'
    s4 = stars(p4) or 'NS'
    if s3 == s4: return f'stable ({s4})'
    sig_order = {'NS': 0, '*': 1, '**': 2, '***': 3}
    if sig_order.get(s4, 0) > sig_order.get(s3, 0): return f'↑ {s3}→{s4}'
    return f'↓ {s3}→{s4}'


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1: H1 — Benchmark regressions
# ══════════════════════════════════════════════════════════════════════════════

H1_SPECS = [
    'Solow_full', 'Malmquist_full',
    'Solow_pars', 'Malmquist_pars',
    'Solow_lag1', 'Malmquist_lag1',
    'Solow_stock', 'Malmquist_stock',
]
H1_ESTIMATORS = ['OLS', 'FE', 'RE', 'CCEP', 'CCEFE']

rows_h1 = []
for spec in H1_SPECS:
    for est in H1_ESTIMATORS:
        r3 = v3.get(spec, {}).get(est)
        r4 = v4.get(spec, {}).get(est)
        if r3 is None and r4 is None:
            continue

        def _get(r, k): return r.get(k) if r else None

        # coef and p are dicts keyed by variable name; extract AI key
        ai_key = 'LN_AI_Patent_Stock' if 'stock' in spec.lower() else \
                 'LN_AI_Patents_L1'   if 'lag1'  in spec.lower() else \
                 'LN_AI_Patents_L2'   if 'lag2'  in spec.lower() else \
                 'LN_AI_Patents'
        coef3 = _get(r3, 'coef'); p_d3 = _get(r3, 'p')
        coef4 = _get(r4, 'coef'); p_d4 = _get(r4, 'p')
        b3 = coef3.get(ai_key) if isinstance(coef3, dict) else coef3
        p3 = p_d3.get(ai_key)  if isinstance(p_d3,  dict) else p_d3
        b4 = coef4.get(ai_key) if isinstance(coef4, dict) else coef4
        p4 = p_d4.get(ai_key)  if isinstance(p_d4,  dict) else p_d4
        n3 = _get(r3, 'obs'); r2_3 = _get(r3, 'r2')
        n4 = _get(r4, 'obs'); r2_4 = _get(r4, 'r2')

        rows_h1.append({
            'Specification':    spec,
            'Estimator':        est,
            # v3
            'v3 β(AI)':         fmt_coef(b3, p3),
            'v3 p-value':       fmt_p(p3),
            'v3 N':             int(n3) if n3 else '—',
            'v3 R²':            f'{r2_3:.3f}' if r2_3 is not None else '—',
            # v4
            'v4 β(AI)':         fmt_coef(b4, p4),
            'v4 p-value':       fmt_p(p4),
            'v4 N':             int(n4) if n4 else '—',
            'v4 R²':            f'{r2_4:.3f}' if r2_4 is not None else '—',
            # delta
            'Δ β':              delta_str(b3, b4),
            'Δ Significance':   delta_sig(p3, p4),
        })

df_h1 = pd.DataFrame(rows_h1)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2: H2 — Mediation
# ══════════════════════════════════════════════════════════════════════════════

rows_h2 = []
for dv in ['Solow', 'Malmquist']:
    for est in ['FE', 'RE']:
        r3 = v3.get('H2_mediation', {}).get(dv, {}).get(est, {})
        r4 = v4.get('H2_mediation', {}).get(dv, {}).get(est, {})

        def g3(k): return r3.get(k)
        def g4(k): return r4.get(k)

        rows_h2.append({
            'DV':                       dv,
            'Estimator':                est,
            # Total effect c
            'v3 c (total, β)':          fmt_coef(g3('step1_c'),   g3('step1_p')),
            'v4 c (total, β)':          fmt_coef(g4('step1_c'),   g4('step1_p')),
            # a-path AI→HC
            'v3 a-path (AI→HC, β)':     fmt_coef(g3('step2_a'),   g3('step2_p')),
            'v4 a-path (AI→HC, β)':     fmt_coef(g4('step2_a'),   g4('step2_p')),
            # b-path HC→TFP
            'v3 b-path (HC→TFP, β)':    fmt_coef(g3('step3_b'),   g3('step3_b_p')),
            'v4 b-path (HC→TFP, β)':    fmt_coef(g4('step3_b'),   g4('step3_b_p')),
            # Direct c'
            "v3 c' (direct, β)":        fmt_coef(g3('step3_c_prime'), g3('step3_c_prime_p')),
            "v4 c' (direct, β)":        fmt_coef(g4('step3_c_prime'), g4('step3_c_prime_p')),
            # Indirect ab
            'v3 Indirect (a×b)':        f"{g3('indirect_ab'):.4f}" if g3('indirect_ab') is not None else '—',
            'v4 Indirect (a×b)':        f"{g4('indirect_ab'):.4f}" if g4('indirect_ab') is not None else '—',
            # Mediation %
            'v3 Mediation %':           f"{g3('mediation_pct'):.1f}%" if g3('mediation_pct') is not None else '—',
            'v4 Mediation %':           f"{g4('mediation_pct'):.1f}%" if g4('mediation_pct') is not None else '—',
            'Δ Mediation %':            (f"{g4('mediation_pct') - g3('mediation_pct'):+.1f}pp"
                                         if g3('mediation_pct') and g4('mediation_pct') else '—'),
        })

df_h2 = pd.DataFrame(rows_h2)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3: H3 — Financial moderation  /  H4 — Digital moderation
# ══════════════════════════════════════════════════════════════════════════════

def build_moderation_df(key3, key4, label):
    rows = []
    for dv in ['Solow', 'Malmquist']:
        for est in ['FE', 'RE']:
            r3 = v3.get(key3, {}).get(dv, {}).get(est, {})
            r4 = v4.get(key4, {}).get(dv, {}).get(est, {})

            def g3(k): return r3.get(k)
            def g4(k): return r4.get(k)

            rows.append({
                'DV':                       dv,
                'Estimator':                est,
                'v3 β(AI)':                 fmt_coef(g3('beta_AI'),          g3('p_AI')),
                'v3 β(AI×MOD)':             fmt_coef(g3('beta_interaction'),  g3('p_interaction')),
                'v3 p(AI×MOD)':             fmt_p(g3('p_interaction')),
                'v3 N':                     int(g3('obs')) if g3('obs') else '—',
                'v4 β(AI)':                 fmt_coef(g4('beta_AI'),          g4('p_AI')),
                'v4 β(AI×MOD)':             fmt_coef(g4('beta_interaction'),  g4('p_interaction')),
                'v4 p(AI×MOD)':             fmt_p(g4('p_interaction')),
                'v4 N':                     int(g4('obs')) if g4('obs') else '—',
                'Δ β(AI×MOD)':              delta_str(g3('beta_interaction'), g4('beta_interaction')),
                'Δ Significance (AI×MOD)':  delta_sig(g3('p_interaction'),    g4('p_interaction')),
            })
    return pd.DataFrame(rows)

df_h3 = build_moderation_df('H3_financial_moderation', 'H3_financial_moderation', 'H3 Financial Moderation')
df_h4 = build_moderation_df('H4_digital_moderation',   'H4_digital_moderation',   'H4 Digital Moderation')

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4: H5 — Panel quantile regression
# ══════════════════════════════════════════════════════════════════════════════

TAUS = ['q10', 'q25', 'q50', 'q75', 'q90']
TAU_LABELS = {'q10': 'τ=0.10', 'q25': 'τ=0.25', 'q50': 'τ=0.50',
              'q75': 'τ=0.75', 'q90': 'τ=0.90'}

rows_h5 = []
for dv in ['Solow', 'Malmquist']:
    for q in TAUS:
        r3 = v3.get('H5_quantile', {}).get(dv, {}).get(q, {})
        r4 = v4.get('H5_quantile', {}).get(dv, {}).get(q, {})

        ai_key = 'LN_AI_Patents'
        b3 = r3.get('coef', {}).get(ai_key)
        p3 = r3.get('p',    {}).get(ai_key)
        b4 = r4.get('coef', {}).get(ai_key)
        p4 = r4.get('p',    {}).get(ai_key)
        n3 = r3.get('obs'); n4 = r4.get('obs')

        rows_h5.append({
            'DV':               dv,
            'Quantile':         TAU_LABELS[q],
            'v3 β(AI)':         fmt_coef(b3, p3),
            'v3 p-value':       fmt_p(p3),
            'v3 N':             int(n3) if n3 else '—',
            'v4 β(AI)':         fmt_coef(b4, p4),
            'v4 p-value':       fmt_p(p4),
            'v4 N':             int(n4) if n4 else '—',
            'Δ β':              delta_str(b3, b4),
            'Δ Significance':   delta_sig(p3, p4),
        })

df_h5 = pd.DataFrame(rows_h5)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5: Pesaran CD test
# ══════════════════════════════════════════════════════════════════════════════

rows_cd = []
for series in ['Solow TFP', 'Malmquist TFP']:
    r3 = v3.get('pesaran_cd', {}).get(series, {})
    r4 = v4.get('pesaran_cd', {}).get(series, {})
    rows_cd.append({
        'Series':       series,
        'v3 CD stat':   f"{r3.get('CD', float('nan')):.3f}" if r3 else '—',
        'v3 p-value':   fmt_p(r3.get('p')),
        'v3 Decision':  'Reject H₀ (CD)' if r3.get('p', 1) < 0.05 else 'Cannot reject H₀',
        'v4 CD stat':   f"{r4.get('CD', float('nan')):.3f}" if r4 else '—',
        'v4 p-value':   fmt_p(r4.get('p')),
        'v4 Decision':  'Reject H₀ (CD)' if r4.get('p', 1) < 0.05 else 'Cannot reject H₀',
        'Δ Decision':   (
            '✓ CD resolved' if (r3.get('p', 1) < 0.05 and r4.get('p', 1) >= 0.05)
            else '✗ CD introduced' if (r3.get('p', 1) >= 0.05 and r4.get('p', 1) < 0.05)
            else 'Stable'
        ),
    })

df_cd = pd.DataFrame(rows_cd)

# ══════════════════════════════════════════════════════════════════════════════
# WRITE EXCEL — one sheet per section
# ══════════════════════════════════════════════════════════════════════════════

OUT.parent.mkdir(parents=True, exist_ok=True)

with pd.ExcelWriter(OUT, engine='openpyxl') as writer:

    # Helper: write df with a header block and return next row
    def write_sheet(df, sheet_name, title, notes=None):
        df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
        ws = writer.sheets[sheet_name]
        ws.cell(1, 1, title)
        ws.cell(2, 1, f'v3 = with imputation (N full=231) | v4 = no imputation | *** p<0.01, ** p<0.05, * p<0.10')
        if notes:
            last_row = len(df) + 5
            for i, n in enumerate(notes):
                ws.cell(last_row + i, 1, n)

    write_sheet(
        df_h1, 'H1_Benchmark',
        'H1 — Benchmark Regressions: AI Patents → TFP (v3 vs v4 No-Imputation)',
        notes=[
            'Note: Hausman test uniformly favours FE over RE (χ² p<0.001 for all Solow specs).',
            'Stock = cumulative patent stock (LN_AI_Patent_Stock). Lag1 = one-year lagged flow.',
            'CCEP/CCEFE estimators not available for Lag and Stock robustness specs.',
        ]
    )
    write_sheet(
        df_h2, 'H2_Mediation',
        'H2 — Baron & Kenny Mediation: AI → HC → TFP (v3 vs v4)',
        notes=['Note: a-path = AI→HC; b-path = HC→TFP; Indirect = a×b; Mediation % = (a×b)/c×100.']
    )
    write_sheet(
        df_h3, 'H3_Financial_Moderation',
        'H3 — Financial Development Moderation: AI × FIN_credit_private → TFP (v3 vs v4)'
    )
    write_sheet(
        df_h4, 'H4_Digital_Moderation',
        'H4 — Digital Infrastructure Moderation: AI × INF_internet → TFP (v3 vs v4)'
    )
    write_sheet(
        df_h5, 'H5_Quantile',
        'H5 — Panel Quantile Regression: AI Patents → TFP at τ=0.10–0.90 (v3 vs v4)',
        notes=['Note: Parsimonious spec; DV = ln_TFP (Solow) or TFP_Change (Malmquist).']
    )
    write_sheet(
        df_cd, 'CD_Test',
        'Pesaran (2004) Cross-Sectional Dependence Test (v3 vs v4)',
        notes=['Note: H₀ = cross-sectional independence; rejection → CCEP/CCEFE preferred over FE.']
    )

    # ── Auto-fit column widths ─────────────────────────────────────────────
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT   = Font(bold=True, size=12, color="1F3864")
    NOTE_FONT    = Font(italic=True, size=9, color="595959")
    V3_FILL      = PatternFill("solid", fgColor="EBF3FB")   # light blue
    V4_FILL      = PatternFill("solid", fgColor="E9F7EF")   # light green
    DELTA_FILL   = PatternFill("solid", fgColor="FEF9E7")   # light yellow
    ALT_FILL     = PatternFill("solid", fgColor="F8F8F8")
    thin         = Side(style='thin', color='CCCCCC')
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

    SHEET_DEFS = {
        'H1_Benchmark':           df_h1,
        'H2_Mediation':           df_h2,
        'H3_Financial_Moderation':df_h3,
        'H4_Digital_Moderation':  df_h4,
        'H5_Quantile':            df_h5,
        'CD_Test':                df_cd,
    }

    for sname, df in SHEET_DEFS.items():
        ws = writer.sheets[sname]
        # Title row
        ws['A1'].font = TITLE_FONT
        ws['A2'].font = NOTE_FONT
        # Header row (row 3)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(3, col_idx)
            cell.font  = HEADER_FONT
            cell.fill  = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = BORDER
            # Colour-code header by column group
            lower = col_name.lower()
            if lower.startswith('v3'):
                cell.fill = PatternFill("solid", fgColor="2E75B6")
            elif lower.startswith('v4'):
                cell.fill = PatternFill("solid", fgColor="375623") if 'v4' in lower else HEADER_FILL
                cell.fill = PatternFill("solid", fgColor="1E6B31")
            elif lower.startswith('δ') or lower.startswith('delta') or 'Δ' in col_name:
                cell.fill = PatternFill("solid", fgColor="7D3C98")
            else:
                cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        # Data rows
        for row_idx in range(len(df)):
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row_idx + 4, col_idx)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal='center')
                # Alternating row fill
                base_fill = ALT_FILL if row_idx % 2 == 1 else PatternFill()
                lower = col_name.lower()
                if lower.startswith('v3'):
                    cell.fill = V3_FILL
                elif lower.startswith('v4'):
                    cell.fill = V4_FILL
                elif 'Δ' in col_name or lower.startswith('δ'):
                    cell.fill = DELTA_FILL
                elif row_idx % 2 == 1:
                    cell.fill = ALT_FILL

                # Bold significance stars
                val = str(cell.value or '')
                if val.endswith('***'):
                    cell.font = Font(bold=True, color='C0392B')
                elif val.endswith('**'):
                    cell.font = Font(bold=True, color='884EA0')
                elif val.endswith('*'):
                    cell.font = Font(bold=True, color='1A5276')
                # Highlight improvements in Δ Significance
                if ('Δ Significance' in col_name or 'Δ Decision' in col_name):
                    if '↑' in val:
                        cell.fill = PatternFill("solid", fgColor="D5F5E3")
                        cell.font = Font(bold=True, color='1E6B31')
                    elif '↓' in val:
                        cell.fill = PatternFill("solid", fgColor="FADBD8")
                        cell.font = Font(bold=True, color='922B21')
                    elif 'resolved' in val.lower():
                        cell.fill = PatternFill("solid", fgColor="D5F5E3")
                        cell.font = Font(bold=True, color='1E6B31')

        # Auto-fit columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(len(str(col_name)),
                         max((len(str(ws.cell(r, col_idx).value or ''))
                              for r in range(4, len(df) + 4)), default=0))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 28)
        # Freeze panes
        ws.freeze_panes = ws.cell(4, 3)

print(f"Saved → {OUT}")
print(f"\nSheets written:")
for s in ['H1_Benchmark', 'H2_Mediation', 'H3_Financial_Moderation',
          'H4_Digital_Moderation', 'H5_Quantile', 'CD_Test']:
    print(f"  • {s}")
