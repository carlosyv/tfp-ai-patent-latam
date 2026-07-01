#!/usr/bin/env python3
"""
build_comparison_v4_table.py
=============================
Side-by-side comparison: v4 (with imputation) vs v4-noimput
Covers H1 benchmark, H2 mediation, H3 institutional, H4 mobile,
H4r broadband robustness, H5 quantile, Pesaran CD test.
"""

import json
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE  = Path(__file__).resolve().parent
V4    = BASE / 'output/results/benchmark_dissertation_v4/regression_results.json'
V4NI  = BASE / 'output/results/benchmark_dissertation_v4_noimput/regression_results.json'
OUT   = BASE / 'output/results/comparison_v4_vs_v4noimput.xlsx'

with open(V4)   as f: v4   = json.load(f)
with open(V4NI) as f: v4ni = json.load(f)

# ── Helpers ────────────────────────────────────────────────────────────────────
def stars(p):
    if p is None or (isinstance(p, float) and np.isnan(p)): return ''
    if p < 0.01:  return '***'
    if p < 0.05:  return '**'
    if p < 0.10:  return '*'
    return ''

def fmt_coef(b, p):
    if b is None or (isinstance(b, float) and np.isnan(b)): return '—'
    return f'{b:.4f}{stars(p)}'

def fmt_p(p):
    if p is None or (isinstance(p, float) and np.isnan(p)): return '—'
    return f'{p:.3f}'

def delta_str(b_base, b_new):
    if b_base is None or b_new is None: return '—'
    d = b_new - b_base
    if abs(d) < 1e-6: return '≈'
    return f'{"↑" if d > 0 else "↓"} {d:+.4f}'

def delta_sig(p_base, p_new):
    s_base = stars(p_base) or 'NS'
    s_new  = stars(p_new)  or 'NS'
    if s_base == s_new: return f'stable ({s_new})'
    sig_order = {'NS': 0, '*': 1, '**': 2, '***': 3}
    if sig_order.get(s_new, 0) > sig_order.get(s_base, 0):
        return f'↑ {s_base}→{s_new}'
    return f'↓ {s_base}→{s_new}'

# ══════════════════════════════════════════════════════════════════════════════
# H1 — Benchmark regressions
# ══════════════════════════════════════════════════════════════════════════════
H1_SPECS = [
    'Solow_full', 'Malmquist_full',
    'Solow_pars', 'Malmquist_pars',
    'Solow_lag1', 'Malmquist_lag1',
    'Solow_stock', 'Malmquist_stock',
]
H1_ESTIMATORS = ['OLS', 'FE', 'RE', 'CCEP', 'CCEFE']

rows_h1 = []
for spec in H1_SPECS:
    for est in H1_ESTIMATORS:
        r_base = v4.get(spec,  {}).get(est)
        r_new  = v4ni.get(spec, {}).get(est)
        if r_base is None and r_new is None:
            continue

        def _get(r, k): return r.get(k) if r else None

        ai_key = ('LN_AI_Patent_Stock' if 'stock' in spec.lower() else
                  'LN_AI_Patents_L1'   if 'lag1'  in spec.lower() else
                  'LN_AI_Patents_L2'   if 'lag2'  in spec.lower() else
                  'LN_AI_Patents')

        def extract(r):
            coef = _get(r, 'coef'); p_d = _get(r, 'p')
            b = coef.get(ai_key) if isinstance(coef, dict) else coef
            p = p_d.get(ai_key)  if isinstance(p_d,  dict) else p_d
            n  = _get(r, 'obs'); r2 = _get(r, 'r2')
            return b, p, n, r2

        b_base, p_base, n_base, r2_base = extract(r_base)
        b_new,  p_new,  n_new,  r2_new  = extract(r_new)

        rows_h1.append({
            'Specification':  spec,
            'Estimator':      est,
            'v4 β(AI)':       fmt_coef(b_base, p_base),
            'v4 p-value':     fmt_p(p_base),
            'v4 N':           int(n_base) if n_base else '—',
            'v4 R²':          f'{r2_base:.3f}' if r2_base is not None else '—',
            'v4-NI β(AI)':    fmt_coef(b_new,  p_new),
            'v4-NI p-value':  fmt_p(p_new),
            'v4-NI N':        int(n_new)  if n_new  else '—',
            'v4-NI R²':       f'{r2_new:.3f}'  if r2_new  is not None else '—',
            'Δ β':            delta_str(b_base, b_new),
            'Δ Significance': delta_sig(p_base, p_new),
        })

df_h1 = pd.DataFrame(rows_h1)

# ══════════════════════════════════════════════════════════════════════════════
# H2 — Mediation
# ══════════════════════════════════════════════════════════════════════════════
rows_h2 = []
for dv in ['Solow', 'Malmquist']:
    for est in ['FE', 'RE']:
        r_base = v4.get('H2_mediation',  {}).get(dv, {}).get(est, {})
        r_new  = v4ni.get('H2_mediation', {}).get(dv, {}).get(est, {})
        def gb(k): return r_base.get(k)
        def gn(k): return r_new.get(k)
        rows_h2.append({
            'DV': dv, 'Estimator': est,
            'v4 c (total)':       fmt_coef(gb('step1_c'),        gb('step1_p')),
            'v4-NI c (total)':    fmt_coef(gn('step1_c'),        gn('step1_p')),
            'v4 a-path (AI→HC)':  fmt_coef(gb('step2_a'),        gb('step2_p')),
            'v4-NI a-path':       fmt_coef(gn('step2_a'),        gn('step2_p')),
            'v4 b-path (HC→TFP)': fmt_coef(gb('step3_b'),        gb('step3_b_p')),
            'v4-NI b-path':       fmt_coef(gn('step3_b'),        gn('step3_b_p')),
            "v4 c' (direct)":     fmt_coef(gb('step3_c_prime'),  gb('step3_c_prime_p')),
            "v4-NI c' (direct)":  fmt_coef(gn('step3_c_prime'),  gn('step3_c_prime_p')),
            'v4 Indirect (a×b)':  f"{gb('indirect_ab'):.4f}" if gb('indirect_ab') is not None else '—',
            'v4-NI Indirect':     f"{gn('indirect_ab'):.4f}" if gn('indirect_ab') is not None else '—',
            'v4 Mediation %':     f"{gb('mediation_pct'):.1f}%" if gb('mediation_pct') is not None else '—',
            'v4-NI Mediation %':  f"{gn('mediation_pct'):.1f}%" if gn('mediation_pct') is not None else '—',
            'Δ Mediation %':      (f"{gn('mediation_pct') - gb('mediation_pct'):+.1f}pp"
                                   if gb('mediation_pct') and gn('mediation_pct') else '—'),
        })
df_h2 = pd.DataFrame(rows_h2)

# ══════════════════════════════════════════════════════════════════════════════
# H3/H4/H4r — Moderation
# ══════════════════════════════════════════════════════════════════════════════
def build_mod_df(json_key, int_label):
    rows = []
    for dv in ['Solow', 'Malmquist']:
        for est in ['FE', 'RE']:
            rb = v4.get(json_key,  {}).get(dv, {}).get(est, {})
            rn = v4ni.get(json_key, {}).get(dv, {}).get(est, {})
            def gb(k): return rb.get(k)
            def gn(k): return rn.get(k)
            rows.append({
                'DV': dv, 'Estimator': est,
                f'v4 β(AI)':               fmt_coef(gb('beta_AI'),         gb('p_AI')),
                f'v4 β(AI×{int_label})':   fmt_coef(gb('beta_interaction'), gb('p_interaction')),
                f'v4 p(AI×{int_label})':   fmt_p(gb('p_interaction')),
                'v4 N':                     int(gb('obs')) if gb('obs') else '—',
                f'v4-NI β(AI)':             fmt_coef(gn('beta_AI'),         gn('p_AI')),
                f'v4-NI β(AI×{int_label})': fmt_coef(gn('beta_interaction'), gn('p_interaction')),
                f'v4-NI p(AI×{int_label})': fmt_p(gn('p_interaction')),
                'v4-NI N':                  int(gn('obs')) if gn('obs') else '—',
                f'Δ β(AI×{int_label})':    delta_str(gb('beta_interaction'), gn('beta_interaction')),
                f'Δ Sig (AI×{int_label})': delta_sig(gb('p_interaction'),    gn('p_interaction')),
            })
    return pd.DataFrame(rows)

df_h3   = build_mod_df('H3_institutional_moderation', 'RL')
df_h4   = build_mod_df('H4_digital_moderation',       'MOB')
df_h4r  = build_mod_df('H4r_broadband_moderation',    'BBND')

# ══════════════════════════════════════════════════════════════════════════════
# H5 — Quantile regression
# ══════════════════════════════════════════════════════════════════════════════
TAUS = {'q10': 'τ=0.10', 'q25': 'τ=0.25', 'q50': 'τ=0.50', 'q75': 'τ=0.75', 'q90': 'τ=0.90'}
rows_h5 = []
for dv in ['Solow', 'Malmquist']:
    for q, qlabel in TAUS.items():
        rb = v4.get('H5_quantile',  {}).get(dv, {}).get(q, {})
        rn = v4ni.get('H5_quantile', {}).get(dv, {}).get(q, {})
        ai_key = 'LN_AI_Patents'
        b_base = rb.get('coef', {}).get(ai_key); p_base = rb.get('p', {}).get(ai_key)
        b_new  = rn.get('coef', {}).get(ai_key); p_new  = rn.get('p', {}).get(ai_key)
        rows_h5.append({
            'DV': dv, 'Quantile': qlabel,
            'v4 β(AI)':       fmt_coef(b_base, p_base),
            'v4 p-value':     fmt_p(p_base),
            'v4 N':           int(rb.get('obs')) if rb.get('obs') else '—',
            'v4-NI β(AI)':    fmt_coef(b_new,  p_new),
            'v4-NI p-value':  fmt_p(p_new),
            'v4-NI N':        int(rn.get('obs')) if rn.get('obs') else '—',
            'Δ β':            delta_str(b_base, b_new),
            'Δ Significance': delta_sig(p_base, p_new),
        })
df_h5 = pd.DataFrame(rows_h5)

# ══════════════════════════════════════════════════════════════════════════════
# CD Test
# ══════════════════════════════════════════════════════════════════════════════
rows_cd = []
for series in ['Solow TFP', 'Malmquist TFP']:
    rb = v4.get('pesaran_cd',  {}).get(series, {})
    rn = v4ni.get('pesaran_cd', {}).get(series, {})
    rows_cd.append({
        'Series':        series,
        'v4 CD stat':    f"{rb.get('CD', float('nan')):.3f}" if rb else '—',
        'v4 p-value':    fmt_p(rb.get('p')),
        'v4 Decision':   'Reject H₀ (CD)' if rb.get('p', 1) < 0.05 else 'Cannot reject H₀',
        'v4-NI CD stat': f"{rn.get('CD', float('nan')):.3f}" if rn else '—',
        'v4-NI p-value': fmt_p(rn.get('p')),
        'v4-NI Decision':'Reject H₀ (CD)' if rn.get('p', 1) < 0.05 else 'Cannot reject H₀',
        'Δ Decision':    (
            '✓ CD resolved'   if (rb.get('p', 1) < 0.05  and rn.get('p', 1) >= 0.05) else
            '✗ CD introduced' if (rb.get('p', 1) >= 0.05 and rn.get('p', 1) < 0.05)  else
            'Stable'
        ),
    })
df_cd = pd.DataFrame(rows_cd)

# ══════════════════════════════════════════════════════════════════════════════
# WRITE EXCEL
# ══════════════════════════════════════════════════════════════════════════════
OUT.parent.mkdir(parents=True, exist_ok=True)

SHEETS = [
    (df_h1,  'H1_Benchmark',
     'H1 — Benchmark Regressions: AI Patents → TFP  (v4 Baseline vs v4 No-Imputation)',
     ['v4 = with imputation (N full=231) | v4-NI = no imputation | *** p<0.01, ** p<0.05, * p<0.10',
      'Stock = cumulative patent stock. Lag1 = one-year lagged flow.',
      'Hausman tests uniformly favour FE for Solow; consult JSON for individual χ² statistics.']),
    (df_h2,  'H2_Mediation',
     'H2 — Baron & Kenny Mediation: AI → HC → TFP  (v4 vs v4-NI)',
     ['a-path = AI→HC | b-path = HC→TFP | Indirect = a×b | Mediation % = (a×b)/c × 100']),
    (df_h3,  'H3_Institutional',
     'H3 — Institutional Quality Moderation: AI × Rule-of-Law → TFP  (v4 vs v4-NI)', []),
    (df_h4,  'H4_Mobile',
     'H4 — Digital Moderation (primary): AI × Mobile/100 → TFP  (v4 vs v4-NI)', []),
    (df_h4r, 'H4r_Broadband',
     'H4r — Digital Moderation Robustness: AI × Broadband/100 → TFP  (v4 vs v4-NI)',
     ['H4r is a robustness check for H4; broadband penetration is lower and sparser than mobile in LAC.']),
    (df_h5,  'H5_Quantile',
     'H5 — Panel Quantile Regression: AI Patents → TFP at τ=0.10–0.90  (v4 vs v4-NI)', []),
    (df_cd,  'CD_Test',
     'Pesaran (2004) Cross-Sectional Dependence Test  (v4 vs v4-NI)',
     ['H₀ = cross-sectional independence. Rejection → use CCEP/CCEFE over standard FE/RE.']),
]

thin   = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

with pd.ExcelWriter(OUT, engine='openpyxl') as writer:
    for df, sname, title, notes in SHEETS:
        df.to_excel(writer, sheet_name=sname, startrow=2, index=False)
        ws = writer.sheets[sname]
        ws['A1'].value = title
        ws['A1'].font  = Font(bold=True, size=12, color='1F3864')
        ws['A2'].value = 'v4 = with imputation | v4-NI = no imputation | *** p<0.01  ** p<0.05  * p<0.10'
        ws['A2'].font  = Font(italic=True, size=9, color='595959')

        # Header row styling
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(3, col_idx)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border    = BORDER
            lower = col_name.lower()
            if lower.startswith('v4-ni') or lower.startswith('v4-ni'):
                cell.fill = PatternFill('solid', fgColor='1E6B31')
            elif lower.startswith('v4'):
                cell.fill = PatternFill('solid', fgColor='2E75B6')
            elif 'Δ' in col_name or lower.startswith('δ'):
                cell.fill = PatternFill('solid', fgColor='7D3C98')
            else:
                cell.fill = PatternFill('solid', fgColor='1F3864')
            cell.font = Font(bold=True, color='FFFFFF', size=10)

        # Data rows
        for row_idx in range(len(df)):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row_idx + 4, col_idx)
                cell.border    = BORDER
                cell.alignment = Alignment(horizontal='center')
                lower = col_name.lower()
                if lower.startswith('v4-ni'):
                    cell.fill = PatternFill('solid', fgColor='E9F7EF')
                elif lower.startswith('v4'):
                    cell.fill = PatternFill('solid', fgColor='EBF3FB')
                elif 'Δ' in col_name:
                    cell.fill = PatternFill('solid', fgColor='FEF9E7')
                elif row_idx % 2 == 1:
                    cell.fill = PatternFill('solid', fgColor='F8F8F8')

                val = str(cell.value or '')
                if val.endswith('***'):
                    cell.font = Font(bold=True, color='C0392B')
                elif val.endswith('**'):
                    cell.font = Font(bold=True, color='884EA0')
                elif val.endswith('*'):
                    cell.font = Font(bold=True, color='1A5276')

                if 'Δ Sig' in col_name or 'Δ Decision' in col_name:
                    if '↑' in val:
                        cell.fill = PatternFill('solid', fgColor='D5F5E3')
                        cell.font = Font(bold=True, color='1E6B31')
                    elif '↓' in val:
                        cell.fill = PatternFill('solid', fgColor='FADBD8')
                        cell.font = Font(bold=True, color='922B21')
                    elif 'resolved' in val.lower():
                        cell.fill = PatternFill('solid', fgColor='D5F5E3')
                        cell.font = Font(bold=True, color='1E6B31')

        # Notes
        if notes:
            last = len(df) + 5
            for i, n in enumerate(notes):
                c = ws.cell(last + i, 1, n)
                c.font = Font(italic=True, size=9, color='595959')

        # Auto-fit column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)),
                         max((len(str(ws.cell(r, col_idx).value or ''))
                              for r in range(4, len(df) + 4)), default=0))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)

        ws.freeze_panes = ws.cell(4, 3)

print(f"Saved → {OUT}")
for _, sname, *_ in SHEETS:
    print(f"  • {sname}")
